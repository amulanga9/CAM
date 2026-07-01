"""
Импорт мастер-Excel (cam_master_2026_06_new.xlsx) в SQLite.
Каждая строка любого "объектного" листа -> одна запись в complexes,
полная исходная строка сохраняется в raw_json (на случай если
понадобится поле, которое мы не выносили в отдельную колонку).

Справочники застройщиков/подрядчиков идут в отдельные таблицы.

Запуск:
    cd ~/CAM/admin
    python3 import_master.py
"""
import json
import re
import sqlite3
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent / 'data' / 'cam_master_2026_06_new.xlsx'

# cam_admin.db — только "слепок" из Excel: complexes/developers/contractors.
# Его можно безопасно удалять и пересобирать сколько угодно раз.
#
# cam_manual.db — твои ручные решения (проверки, исключённые "не ЖК",
# правки полей, вручную добавленные объекты без CAM ID). Этот файл НИКОГДА
# не дропается и не пересоздаётся этим скриптом — удалять его нельзя,
# иначе пропадёт вся ручная работа.
DB_PATH = Path(__file__).parent / 'cam_admin.db'
MANUAL_DB_PATH = Path(__file__).parent / 'cam_manual.db'

OBJECT_SHEETS = [
    'Обучение',
    'Строящиеся_сматченные',
    'Строящиеся_несматченные',
    'Все построенные',
    'Плохие кейсы',
    'Серая зона (1-2 года)',
]
REF_SHEETS = {
    'Все застройщики': 'developers',
    'Все подрядчики': 'contractors',
}

# canonical_field -> возможные имена колонок в разных листах (по порядку приоритета)
FIELD_ALIASES = {
    'cam_id':            ['cam_id', 'CAM ЖК ID'],
    'project_name':      ['project_name', 'dom_name'],
    'address':           ['address'],
    'lat':               ['lat'],
    'lng':               ['lng', 'long'],
    'district_name':     ['district_name'],
    'developer_name':    ['developer_name_norm', 'developer_name_original'],
    'developer_inn':     ['developer_inn'],
    'developer_rating':  ['developer_rating'],
    'contractor_name':   ['contractor_name_norm', 'contractor_name_original'],
    'contractor_inn':    ['contractor_inn'],
    'contractor_rating': ['contractor_rating'],
    'created_at':        ['created_at'],
    'deadline':          ['deadline', 'deadline_max', 'deadline_min'],
    'case_status':       ['case_status', 'shaffof_status'],
    'master_status':     ['master_status'],
    'target':            ['target'],
    'is_overdue':        ['is_overdue'],
    'days_overdue':      ['days_overdue'],
    'dom_class':         ['dom_class'],
    'price_per_m2_uzs':  ['price_per_m2_uzs'],
}

# нормализация "грязного" case_status в чистый бакет.
# всё, что не распознано однозначно -> 'unclear' (=требует ручной проверки)
CLEAN_STATUS_MAP = {
    'Topshirilgan': 'delivered',
    "To'xtatilgan": 'stopped',
    'Jarayonda': 'in_progress',
    'Bekor qilingan': 'cancelled',
    'Muzlatilgan': 'frozen',
    'Просрочен': 'unclear',
}


def clean_status(raw):
    if not raw:
        return 'unknown'
    raw = str(raw).strip()
    if raw in CLEAN_STATUS_MAP:
        return CLEAN_STATUS_MAP[raw]
    return 'unclear'


def get_field(row, header_idx, canonical):
    for alias in FIELD_ALIASES[canonical]:
        i = header_idx.get(alias)
        if i is not None and i < len(row) and row[i] not in (None, ''):
            return row[i]
    return None


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def to_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


CORPUS_LABEL_RE = re.compile(
    r'\(([^()]*(?:blok|bosqich|korpus|очеред|корпус|блок|этап)[^()]*)\)',
    re.IGNORECASE,
)


def extract_corpus_label(project_name):
    """Достаёт пометку корпуса/очереди/блока из текста в скобках в названии
    проекта (напр. "...(4-blok 2-bosqich A va B korpus)"), чтобы подписать
    фазу понятным человеку текстом вместо безликого номера листа."""
    if not project_name:
        return None
    m = CORPUS_LABEL_RE.search(str(project_name))
    return m.group(1).strip() if m else None


def migrate_complexes(conn):
    """Добавляет недавно введённые колонки в существующую complexes,
    если БД создана старой версией схемы (без полной пересборки из Excel)."""
    cols = {r[1] for r in conn.execute('PRAGMA table_info(complexes)')}
    if not cols:
        return  # таблицы ещё нет — её создаст build_schema/ensure_complexes_table
    for col in ('brand_name', 'delivered_year', 'holding_name'):
        if col not in cols:
            conn.execute(f'ALTER TABLE complexes ADD COLUMN {col} TEXT')
    if 'brand_status' not in cols:
        conn.execute('ALTER TABLE complexes ADD COLUMN brand_status TEXT')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS phases (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            cam_id            TEXT,
            source_sheet      TEXT,
            corpus_label      TEXT,
            project_name      TEXT,
            address           TEXT,
            case_status_raw   TEXT,
            case_status_clean TEXT,
            master_status     TEXT,
            target            INTEGER,
            created_at        TEXT,
            deadline          TEXT,
            delivery_date_fact TEXT,
            delay_days_fact   TEXT,
            days_overdue      INTEGER
        )
    ''')


def build_schema(conn):
    # complexes/developers/contractors пересобираются с нуля при каждом импорте,
    # т.к. их источник — Excel. Ручные данные живут в отдельном файле
    # cam_manual.db (см. build_manual_schema) и сюда не попадают.
    conn.executescript('''
    DROP TABLE IF EXISTS complexes;
    DROP TABLE IF EXISTS developers;
    DROP TABLE IF EXISTS contractors;
    DROP TABLE IF EXISTS phases;

    CREATE TABLE complexes (
        cam_id            TEXT PRIMARY KEY,
        source_sheet      TEXT,
        project_name      TEXT,
        address           TEXT,
        lat               REAL,
        lng               REAL,
        district_name     TEXT,
        developer_name    TEXT,
        developer_inn     TEXT,
        developer_rating  TEXT,
        contractor_name   TEXT,
        contractor_inn    TEXT,
        contractor_rating TEXT,
        created_at        TEXT,
        deadline          TEXT,
        case_status_raw   TEXT,
        case_status_clean TEXT,
        master_status     TEXT,
        target            INTEGER,
        is_overdue        INTEGER,
        days_overdue      INTEGER,
        dom_class         TEXT,
        price_per_m2_uzs  REAL,
        listing_url       TEXT,
        brand_name        TEXT,
        brand_status      TEXT,
        delivered_year    TEXT,
        holding_name      TEXT,
        needs_review      INTEGER DEFAULT 0,
        raw_json          TEXT
    );

    -- одна строка ИЗ ЛЮБОГО листа на каждую встреченную запись объекта.
    -- для одиночных объектов это всегда 1 строка (совпадает с complexes);
    -- для group-ID (несколько фаз/корпусов под одним cam_id) — несколько,
    -- чтобы при пересборке complexes (берёт только первую/худшую) не терялась
    -- видимость на остальные фазы.
    CREATE TABLE phases (
        id                INTEGER PRIMARY KEY AUTOINCREMENT,
        cam_id            TEXT,
        source_sheet      TEXT,
        corpus_label      TEXT,
        project_name      TEXT,
        address           TEXT,
        case_status_raw   TEXT,
        case_status_clean TEXT,
        master_status     TEXT,
        target            INTEGER,
        created_at        TEXT,
        deadline          TEXT,
        delivery_date_fact TEXT,
        delay_days_fact   TEXT,
        days_overdue      INTEGER
    );

    CREATE TABLE developers (
        name_norm     TEXT PRIMARY KEY,
        name_original TEXT,
        inn           TEXT,
        rating        TEXT,
        objects_count INTEGER,
        bad_pct       REAL,
        overdue_pct   REAL,
        complexes_count INTEGER
    );

    CREATE TABLE contractors (
        name_norm     TEXT PRIMARY KEY,
        name_original TEXT,
        inn           TEXT,
        rating        TEXT,
        objects_count INTEGER,
        bad_pct       REAL,
        overdue_pct   REAL,
        complexes_count INTEGER
    );
    ''')


def attach_manual_db(conn):
    conn.execute(f"ATTACH DATABASE '{MANUAL_DB_PATH}' AS manual")


def build_manual_schema(conn):
    """cam_manual.db — никогда не дропается, только CREATE IF NOT EXISTS."""
    conn.executescript('''
    CREATE TABLE IF NOT EXISTS manual.reviews (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        cam_id      TEXT NOT NULL,
        reviewed_at TEXT NOT NULL,
        verdict     TEXT NOT NULL,   -- delivered | building | frozen | cancelled
        comment     TEXT
    );

    -- объекты-мусор (не ЖК), которые больше не должны попадать ни в импорт,
    -- ни в будущий парсинг
    CREATE TABLE IF NOT EXISTS manual.excluded_ids (
        cam_id      TEXT PRIMARY KEY,
        reason      TEXT,
        excluded_at TEXT
    );

    -- ручные правки полей карточки (дозаполнение/исправление недостающих
    -- данных). Переживают реимпорт и переприменяются к complexes после
    -- пересборки из Excel.
    CREATE TABLE IF NOT EXISTS manual.overrides (
        cam_id            TEXT PRIMARY KEY,
        project_name      TEXT,
        address           TEXT,
        district_name     TEXT,
        lat               REAL,
        lng               REAL,
        developer_name    TEXT,
        developer_inn     TEXT,
        developer_rating  TEXT,
        contractor_name   TEXT,
        contractor_inn    TEXT,
        contractor_rating TEXT,
        deadline          TEXT,
        dom_class         TEXT,
        price_per_m2_uzs  REAL,
        listing_url       TEXT,
        brand_name        TEXT,
        brand_status      TEXT,
        delivered_year    TEXT,
        holding_name      TEXT,
        updated_at        TEXT
    );

    -- объекты, добавленные вручную через админку (ещё не получили CAM ID
    -- из основного пайплайна), идентифицируются по адресу
    CREATE TABLE IF NOT EXISTS manual.manual_complexes (
        cam_id       TEXT PRIMARY KEY,
        project_name TEXT,
        address      TEXT NOT NULL,
        district_name TEXT,
        lat          REAL,
        lng          REAL,
        note         TEXT,
        created_at   TEXT
    );

    -- промежуточный список ребрендов/перезапусков (объект продолжает
    -- строиться под новым застройщиком/названием при том же юр. адресе).
    -- НИКАКОГО нового CAM ID здесь не создаётся — просто копим записи,
    -- matched_cam_id заполняется позже (вручную или автоматическим
    -- матчингом по адресу/координатам/названию с другими источниками,
    -- напр. domtut), как только найдётся соответствующая карточка.
    CREATE TABLE IF NOT EXISTS manual.rebrands (
        id                      INTEGER PRIMARY KEY AUTOINCREMENT,
        cam_id                  TEXT NOT NULL,
        rebrand_name            TEXT,
        rebrand_developer_name  TEXT,
        rebrand_developer_inn   TEXT,
        rebrand_contractor_name TEXT,
        rebrand_contractor_inn  TEXT,
        old_developer_remained  TEXT,
        old_contractor_remained TEXT,
        transition_date         TEXT,
        reason                  TEXT,
        matched_cam_id          TEXT,
        created_at              TEXT
    );

    -- метка переноса сроков, отдельная от вердикта по статусу: у объекта
    -- может быть статус "строится"/"заморожен" и при этом отдельно отмечен
    -- перенос дедлайна. delay_type: unconfirmed (есть только сами даты,
    -- без документа) | confirmed (есть proof_url+proof_type — официальный
    -- источник про САМ объект, не догадка). Сервер сам понижает confirmed
    -- до unconfirmed, если proof_url/proof_type не заполнены.
    CREATE TABLE IF NOT EXISTS manual.delay_flags (
        cam_id            TEXT PRIMARY KEY,
        delay_type        TEXT NOT NULL,   -- unconfirmed | confirmed
        original_deadline TEXT,
        current_deadline  TEXT,
        proof_url         TEXT,
        proof_type        TEXT,            -- gasn_closed | suspended_registry | court | news_named
        comment           TEXT,
        updated_at        TEXT,
        permit_date            TEXT,   -- Qurilishga ruxsat berilgan sana — дата выдачи разрешения на строительство
        smr_registration_date  TEXT,   -- дата из выписки о регистрации начала СМР (может совпадать с датой генерации документа, а не события)
        application_number     TEXT,   -- Ариза рақами — номер заявления, для сверки документов друг с другом
        doc_hash               TEXT,   -- № выписки (уникальный номер документа на repo.gov.uz) — для повторной проверки/дедупа загрузок
        applicant_name         TEXT,   -- "Ҳужжат берилган" — застройщик/физлицо из самой выписки, для сверки с complexes.developer_name
        tax_id                 TEXT,   -- СТИР (юрлицо) или ЖШШИР (физлицо) из выписки, для сверки с complexes.developer_inn
        portal_deadline        TEXT,   -- endDate с object-info портала — собственный дедлайн nazorat.mc.uz, для сверки с complexes.deadline
        expertise_number       TEXT,   -- реестровый номер заключения экспертизы (с object-info)
        apz_number             TEXT,   -- реестровый номер заключения Архитектурно-градостроительного совета
        portal_fetched_at      TEXT    -- когда последний раз автоматически подтягивали object-info с портала
    );
    ''')
    for col in ('permit_date', 'smr_registration_date', 'application_number',
                'doc_hash', 'applicant_name', 'tax_id',
                'portal_deadline', 'expertise_number', 'apz_number', 'portal_fetched_at'):
        try:
            conn.execute(f'ALTER TABLE manual.delay_flags ADD COLUMN {col} TEXT')
        except sqlite3.OperationalError:
            pass
    conn.execute('''
    CREATE TABLE IF NOT EXISTS manual.complex_proofs (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        cam_id          TEXT NOT NULL,
        proof_type      TEXT,
        proof_doc_number TEXT,
        proof_ariza     TEXT,
        proof_verify_url TEXT,
        proof_extracted TEXT,
        added_at        TEXT
    );
    ''')
    _migrate_manual_overrides(conn)
    _migrate_manual_rebrands(conn)


def _migrate_manual_rebrands(conn):
    """Добавляет колонки в manual.rebrands, если таблица создана старой версией схемы."""
    cols = {r[1] for r in conn.execute("PRAGMA manual.table_info(rebrands)")}
    for col in ('rebrand_contractor_name', 'rebrand_contractor_inn'):
        if col not in cols:
            conn.execute(f'ALTER TABLE manual.rebrands ADD COLUMN {col} TEXT')


def _migrate_manual_overrides(conn):
    """Добавляет колонки в manual.overrides, если БД создана старой версией схемы."""
    cols = {r[1] for r in conn.execute("PRAGMA manual.table_info(overrides)")}
    for col in ('brand_name', 'delivered_year', 'holding_name', 'brand_status'):
        if col not in cols:
            conn.execute(f'ALTER TABLE manual.overrides ADD COLUMN {col} TEXT')


def load_excluded_ids(conn):
    return {r[0] for r in conn.execute('SELECT cam_id FROM manual.excluded_ids')}


def import_objects(conn, wb):
    cur = conn.cursor()
    excluded = load_excluded_ids(conn)
    seen = set()
    total = 0
    skipped_excluded = 0
    for sheet_name in OBJECT_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f'  пропущен лист (не найден): {sheet_name}')
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        header_idx = {h: i for i, h in enumerate(header) if h is not None}

        n = 0
        for row in rows[1:]:
            cam_id = get_field(row, header_idx, 'cam_id')
            if not cam_id:
                continue
            cam_id = str(cam_id)

            case_status_raw = get_field(row, header_idx, 'case_status')
            cs_clean = clean_status(case_status_raw)
            is_overdue = bool(get_field(row, header_idx, 'is_overdue'))
            needs_review = 1 if (cs_clean == 'unclear' or (cs_clean == 'in_progress' and is_overdue)) else 0
            if cam_id in excluded:
                # "не ЖК" — не показываем в очереди проверки, но строку (с ИНН/застройщиком
                # /подрядчиком) сохраняем, чтобы её можно было найти по названию/ИНН позже
                skipped_excluded += 1
                cs_clean = 'excluded'
                needs_review = 0

            raw_dict = {h: row[i] for h, i in header_idx.items()}
            phase_project_name = get_field(row, header_idx, 'project_name')

            # фиксируем каждую встреченную строку как "фазу" — для одиночных
            # объектов это единственная строка, для group-ID (несколько
            # корпусов/очередей под одним cam_id) видно все, даже те, что
            # complexes ниже не возьмёт из-за дедупликации по cam_id.
            # corpus_label вытаскивается из текста в скобках в названии —
            # это и есть человекочитаемая подпись фазы (корпус/очередь/блок)
            cur.execute('''
                INSERT INTO phases (
                    cam_id, source_sheet, corpus_label, project_name, address,
                    case_status_raw, case_status_clean,
                    master_status, target, created_at, deadline,
                    delivery_date_fact, delay_days_fact, days_overdue
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                cam_id, sheet_name,
                extract_corpus_label(phase_project_name),
                str(phase_project_name or ''),
                str(get_field(row, header_idx, 'address') or ''),
                str(case_status_raw or ''), cs_clean,
                get_field(row, header_idx, 'master_status'),
                to_int(get_field(row, header_idx, 'target')),
                str(get_field(row, header_idx, 'created_at') or ''),
                str(get_field(row, header_idx, 'deadline') or ''),
                str(raw_dict.get('delivery_date_fact') or ''),
                str(raw_dict.get('delay_days_fact') or ''),
                to_int(get_field(row, header_idx, 'days_overdue')),
            ))

            if cam_id in seen:
                continue  # одна и та же запись может встречаться на нескольких листах — для complexes берём первую
            seen.add(cam_id)

            cur.execute('''
                INSERT OR IGNORE INTO complexes (
                    cam_id, source_sheet, project_name, address, lat, lng,
                    district_name, developer_name, developer_inn, developer_rating,
                    contractor_name, contractor_inn, contractor_rating,
                    created_at, deadline, case_status_raw, case_status_clean,
                    master_status, target, is_overdue, days_overdue,
                    dom_class, price_per_m2_uzs, needs_review, raw_json
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                cam_id, sheet_name,
                get_field(row, header_idx, 'project_name'),
                get_field(row, header_idx, 'address'),
                to_float(get_field(row, header_idx, 'lat')),
                to_float(get_field(row, header_idx, 'lng')),
                get_field(row, header_idx, 'district_name'),
                get_field(row, header_idx, 'developer_name'),
                str(get_field(row, header_idx, 'developer_inn') or ''),
                get_field(row, header_idx, 'developer_rating'),
                get_field(row, header_idx, 'contractor_name'),
                str(get_field(row, header_idx, 'contractor_inn') or ''),
                get_field(row, header_idx, 'contractor_rating'),
                str(get_field(row, header_idx, 'created_at') or ''),
                str(get_field(row, header_idx, 'deadline') or ''),
                str(case_status_raw or ''),
                cs_clean,
                get_field(row, header_idx, 'master_status'),
                to_int(get_field(row, header_idx, 'target')),
                int(is_overdue),
                to_int(get_field(row, header_idx, 'days_overdue')),
                get_field(row, header_idx, 'dom_class'),
                to_float(get_field(row, header_idx, 'price_per_m2_uzs')),
                needs_review,
                json.dumps(raw_dict, ensure_ascii=False, default=str),
            ))
            n += 1
        print(f'  {sheet_name}: {n} новых записей')
        total += n
    conn.commit()
    print(f'Всего объектов: {total} (пропущено как "не ЖК": {skipped_excluded})')


# поля, которые можно дозаполнить/исправить вручную в карточке объекта
EDITABLE_FIELDS = [
    'project_name', 'address', 'district_name', 'lat', 'lng',
    'developer_name', 'developer_inn', 'developer_rating',
    'contractor_name', 'contractor_inn', 'contractor_rating',
    'deadline', 'dom_class', 'price_per_m2_uzs', 'listing_url',
    'brand_name', 'brand_status', 'delivered_year', 'holding_name',
]


def reapply_overrides(conn):
    """Возвращает ранее введённые вручную поля поверх пересобранных complexes."""
    cur = conn.cursor()
    rows = cur.execute(f'SELECT cam_id, {", ".join(EDITABLE_FIELDS)} FROM manual.overrides').fetchall()
    n = 0
    for row in rows:
        cam_id, values = row[0], row[1:]
        sets = [(f, v) for f, v in zip(EDITABLE_FIELDS, values) if v is not None and v != '']
        if not sets:
            continue
        assignment = ', '.join(f'{f}=?' for f, _ in sets)
        cur.execute(f'UPDATE complexes SET {assignment} WHERE cam_id=?', [v for _, v in sets] + [cam_id])
        n += cur.rowcount
    conn.commit()
    print(f'Применены ранее сохранённые правки полей: {n}')


def reapply_reviews(conn):
    """После пересборки complexes возвращаем ранее принятые решения,
    чтобы повторный импорт не сбрасывал то, что уже проверено вручную."""
    cur = conn.cursor()
    latest = cur.execute('''
        SELECT cam_id, verdict FROM manual.reviews r
        WHERE reviewed_at = (SELECT MAX(reviewed_at) FROM manual.reviews WHERE cam_id = r.cam_id)
    ''').fetchall()
    n = 0
    for cam_id, verdict in latest:
        cur.execute(
            'UPDATE complexes SET case_status_clean=?, needs_review=0 WHERE cam_id=?',
            (verdict, cam_id),
        )
        n += cur.rowcount
    conn.commit()
    print(f'Применены ранее сохранённые решения: {n}')


def merge_manual_complexes(conn):
    cur = conn.cursor()
    rows = cur.execute('SELECT * FROM manual.manual_complexes').fetchall()
    cols = [d[0] for d in cur.description]
    n = 0
    for row in rows:
        d = dict(zip(cols, row))
        cur.execute('''
            INSERT OR IGNORE INTO complexes (
                cam_id, source_sheet, project_name, address, lat, lng,
                district_name, case_status_clean, needs_review, raw_json
            ) VALUES (?,?,?,?,?,?,?,?,?,?)
        ''', (
            d['cam_id'], 'manual', d['project_name'], d['address'],
            d['lat'], d['lng'], d['district_name'], 'unknown', 0,
            json.dumps(d, ensure_ascii=False, default=str),
        ))
        n += cur.rowcount
    conn.commit()
    print(f'Добавлено вручную добавленных объектов: {n}')


def import_ref_sheet(conn, wb, sheet_name, table):
    if sheet_name not in wb.sheetnames:
        print(f'  пропущен лист (не найден): {sheet_name}')
        return
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    cur = conn.cursor()
    n = 0
    for row in rows[1:]:
        name_norm, name_orig, inn, rating, obj_cnt, bad_pct, overdue_pct, cplx_cnt = (list(row) + [None] * 8)[:8]
        if not name_norm:
            continue
        cur.execute(f'''
            INSERT OR REPLACE INTO {table}
            (name_norm, name_original, inn, rating, objects_count, bad_pct, overdue_pct, complexes_count)
            VALUES (?,?,?,?,?,?,?,?)
        ''', (str(name_norm), name_orig, str(inn or ''), rating,
              to_int(obj_cnt), to_float(bad_pct), to_float(overdue_pct), to_int(cplx_cnt)))
        n += 1
    conn.commit()
    print(f'  {sheet_name} -> {table}: {n} записей')


def main():
    print(f'Читаю {XLSX_PATH} ...')
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    conn = sqlite3.connect(DB_PATH)
    attach_manual_db(conn)
    build_manual_schema(conn)
    build_schema(conn)

    print('Импорт объектов:')
    import_objects(conn, wb)

    print('Импорт справочников:')
    for sheet, table in REF_SHEETS.items():
        import_ref_sheet(conn, wb, sheet, table)

    merge_manual_complexes(conn)
    reapply_reviews(conn)
    reapply_overrides(conn)

    cur = conn.execute('SELECT COUNT(*) FROM complexes WHERE needs_review=1')
    print(f'\nТребуют проверки (needs_review=1): {cur.fetchone()[0]}')
    conn.close()
    print(f'\nГотово: {DB_PATH} (ручные данные в {MANUAL_DB_PATH})')


if __name__ == '__main__':
    main()
